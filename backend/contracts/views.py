import io
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from accounts.permissions import RoleSectionPermission
from accounts.mixins import SafeDestroyMixin
from .models import Customer, Contract, PaymentScheduleItem, ContractFile, Comment
from .serializers import (CustomerSerializer, ContractSerializer, ContractDetailSerializer,
                          PaymentScheduleItemSerializer, ContractFileSerializer, CommentSerializer)

EXPORT_HEADERS = ["number", "customer", "title", "status", "amount",
                  "signed_date", "deadline", "specification"]


def parse_amount_cell(value):
    """Сумма из ячейки. Люди копируют её из отчётов вместе с пробелами
    и знаком тенге: «450 000 ₸». Раньше такая строка роняла весь договор."""
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return value
    text = (str(value).replace("\xa0", "").replace(" ", "")
            .replace("₸", "").replace("тг", "").replace(",", "."))
    try:
        return float(text)
    except ValueError:
        return 0


CARRY_OVER_NOTE = "Перенос из прежнего учёта"


def carry_over_payment(contract, raw_paid):
    """Перенести из файла то, что по договору уже оплачено.

    Выгрузка колонку paid_amount пишет, а импорт её раньше не читал — при
    переносе истории оплаченный договор выглядел неоплаченным, и воронка
    считала уже полученные деньги ожидаемыми.

    Оплаты живут в графике платежей, поэтому заводим одну строку на всю
    сумму договора с отметкой о переносе. Повторная загрузка того же файла
    её обновляет, а не добавляет вторую: строка узнаётся по этой отметке.
    """
    paid = parse_amount_cell(raw_paid)
    item = contract.payment_schedule.filter(note=CARRY_OVER_NOTE).first()
    if not paid:
        if item:
            item.delete()
        return
    if contract.amount <= 0:
        return
    due = contract.deadline or contract.signed_date or timezone.localdate()
    values = dict(amount=contract.amount, paid_amount=min(paid, contract.amount),
                  due_date=due, paid_date=contract.signed_date or due)
    if item:
        for k, v in values.items():
            setattr(item, k, v)
        item.save()
    else:
        PaymentScheduleItem.objects.create(contract=contract, note=CARRY_OVER_NOTE, **values)


def parse_date_cell(value):
    """Дата из текстовой ячейки. None, если формат непонятен.

    Ячейку с настоящим форматом даты Excel отдаёт как datetime — её разбирать
    не надо. Текстом дату пишут по-разному, и раньше принимался только
    ISO-вид: строка с «15.09.2026» роняла весь договор, а не одно поле.
    """
    text = str(value).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


class CustomerViewSet(SafeDestroyMixin, viewsets.ModelViewSet):
    access_key = "contracts.customers"
    queryset = Customer.objects.all().order_by("name")
    serializer_class = CustomerSerializer
    permission_classes = [RoleSectionPermission]
    section = "contracts"
    search_fields = ["name", "bin_iin", "phone"]


class ContractViewSet(SafeDestroyMixin, viewsets.ModelViewSet):
    access_key = "contracts.contracts"
    queryset = Contract.objects.select_related("customer", "manager")
    permission_classes = [RoleSectionPermission]
    section = "contracts"
    filterset_fields = ["status", "customer", "manager"]
    search_fields = ["number", "title", "customer__name"]

    def get_serializer_class(self):
        return ContractDetailSerializer if self.action == "retrieve" else ContractSerializer

    @action(detail=True, methods=["post"])
    def set_status(self, request, pk=None):
        """Смена статуса строго по цепочке: new → negotiation → in_progress → closed/cancelled."""
        contract = self.get_object()
        new_status = request.data.get("status")
        if new_status not in dict(Contract.Status.choices):
            return Response({"detail": "Неизвестный статус договора."},
                            status=status.HTTP_400_BAD_REQUEST)
        err = contract.transition_error(new_status)
        if err:
            return Response({"detail": err}, status=status.HTTP_400_BAD_REQUEST)
        contract.status = new_status
        contract.save(update_fields=["status", "updated_at"])
        return Response(ContractDetailSerializer(contract).data)

    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        """Выгрузка договоров в Excel для отчётности."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Договоры"
        ws.append(EXPORT_HEADERS + ["paid_amount"])
        for c in self.filter_queryset(self.get_queryset()):
            ws.append([c.number, c.customer.name, c.title, c.status, float(c.amount),
                       str(c.signed_date or ""), str(c.deadline or ""),
                       c.specification, float(c.paid_amount)])
        buf = io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(buf.getvalue(),
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="contracts.xlsx"'
        return resp

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser])
    def import_excel(self, request):
        """Массовая загрузка: колонки number, customer, title, status, amount,
        paid_amount, signed_date, deadline, specification.

        Поле формы carry_over=1 включает режим переноса истории: статус берётся
        из файла как есть, даже если по цепочке такой переход запрещён. Нужно
        при переезде с прежнего учёта — договоры туда попадают уже выполненными,
        и проводить каждый через согласование бессмысленно.
        """
        from openpyxl import load_workbook
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "Файл не передан (поле file)."}, status=400)
        carry_over = str(request.data.get("carry_over", "")).lower() in ("1", "true", "on")
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip().lower() if h else "" for h in rows[0]]
        created, updated, errors = 0, 0, []
        for i, row in enumerate(rows[1:], start=2):
            data = dict(zip(header, row))
            try:
                number = str(data.get("number") or "").strip()
                if not number:
                    continue
                customer, _ = Customer.objects.get_or_create(name=str(data.get("customer") or "Без имени").strip())
                defaults = {
                    "customer": customer,
                    "title": str(data.get("title") or "")[:255],
                    "amount": parse_amount_cell(data.get("amount")),
                    "specification": str(data.get("specification") or ""),
                }
                for f in ("signed_date", "deadline"):
                    v = data.get(f)
                    if isinstance(v, datetime):
                        defaults[f] = v.date()
                    elif isinstance(v, str) and v.strip():
                        d = parse_date_cell(v)
                        if d is None:
                            errors.append(f"строка {i}: дату «{v.strip()}» разобрать не удалось, "
                                          "поле оставлено пустым. Ожидается 2026-09-15 или 15.09.2026.")
                        else:
                            defaults[f] = d

                # Статус подчиняется той же цепочке, что и кнопки в карточке.
                # У нового договора он может быть любым — в систему вносят и те,
                # что давно в работе. У существующего запрещённый переход не
                # применяется молча: строка загружается, статус остаётся прежним,
                # а причина попадает в список замечаний.
                existing = Contract.objects.filter(number=number).first()
                st = str(data.get("status") or "").strip()
                if st in dict(Contract.Status.choices):
                    if existing is None or carry_over:
                        defaults["status"] = st
                    else:
                        err = existing.transition_error(st)
                        if err:
                            errors.append(f"строка {i}: {err} Статус договора не изменён. "
                                          "Если это перенос истории, включите режим переноса.")
                        else:
                            defaults["status"] = st

                contract, was_created = Contract.objects.update_or_create(
                    number=number, defaults=defaults)
                created += was_created
                updated += (not was_created)

                carry_over_payment(contract, data.get("paid_amount"))
            except Exception as e:
                errors.append(f"строка {i}: {e}")
        return Response({"created": created, "updated": updated, "errors": errors})


class PaymentScheduleViewSet(SafeDestroyMixin, viewsets.ModelViewSet):
    access_key = "contracts.schedule"
    queryset = PaymentScheduleItem.objects.select_related("contract")
    serializer_class = PaymentScheduleItemSerializer
    permission_classes = [RoleSectionPermission]
    section = "contracts"
    filterset_fields = ["contract"]

    def _register_cash(self, item, delta):
        """Оплата по графику автоматически попадает в кассу (Cash Flow)."""
        from django.utils import timezone
        from finance.models import CashEntry
        if delta > 0:
            CashEntry.objects.create(
                direction="in", amount=delta,
                date=item.paid_date or timezone.localdate(),
                contract=item.contract,
                description=f"Оплата по графику, договор №{item.contract.number}")

    def perform_create(self, serializer):
        item = serializer.save()
        self._register_cash(item, item.paid_amount)

    def perform_update(self, serializer):
        old_paid = serializer.instance.paid_amount
        item = serializer.save()
        self._register_cash(item, item.paid_amount - old_paid)


class ContractFileViewSet(SafeDestroyMixin, viewsets.ModelViewSet):
    access_key = "contracts.files"
    queryset = ContractFile.objects.all()
    serializer_class = ContractFileSerializer
    permission_classes = [RoleSectionPermission]
    section = "contracts"
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filterset_fields = ["contract", "kind"]

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)


class CommentViewSet(SafeDestroyMixin, viewsets.ModelViewSet):
    access_key = "contracts.comments"
    queryset = Comment.objects.select_related("author", "contract")
    serializer_class = CommentSerializer
    permission_classes = [RoleSectionPermission]
    section = "contracts"
    filterset_fields = ["contract", "importance"]

    def perform_create(self, serializer):
        serializer.save(author=self.request.user)

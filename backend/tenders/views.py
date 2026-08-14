import io
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from accounts.permissions import RoleSectionPermission
from accounts.mixins import SafeDestroyMixin
from .models import Tender, Platform, OwnCompany
from .serializers import TenderSerializer, PlatformSerializer, OwnCompanySerializer


class TenderBase(SafeDestroyMixin, viewsets.ModelViewSet):
    permission_classes = [RoleSectionPermission]
    section = "tenders"


class PlatformViewSet(TenderBase):
    queryset = Platform.objects.all()
    serializer_class = PlatformSerializer


class OwnCompanyViewSet(TenderBase):
    queryset = OwnCompany.objects.all()
    serializer_class = OwnCompanySerializer


class TenderViewSet(TenderBase):
    queryset = Tender.objects.select_related("platform", "own_company", "product",
                                             "manager", "contract")
    serializer_class = TenderSerializer
    filterset_fields = ["status", "platform", "own_company", "manager"]
    search_fields = ["purchase_no", "lot_no", "item_name", "customer_name"]

    @action(detail=True, methods=["post"])
    def set_status(self, request, pk=None):
        """Движение по воронке строго по разрешённым переходам."""
        tender = self.get_object()
        new_status = request.data.get("status")
        allowed = Tender.TRANSITIONS.get(tender.status, set())
        if new_status not in allowed:
            return Response(
                {"detail": f"Переход {tender.status} → {new_status} запрещён. "
                           f"Доступно: {sorted(map(str, allowed))}"}, status=400)
        tender.status = new_status
        if new_status == Tender.Status.SUBMITTED and not tender.submitted_at:
            tender.submitted_at = timezone.localdate()
        tender.save()
        return Response(TenderSerializer(tender).data)

    @action(detail=True, methods=["post"])
    def calc_cost(self, request, pk=None):
        """Подтянуть себестоимость из каталога (BOM + труд + накладные)."""
        tender = self.get_object()
        if not tender.product:
            return Response({"detail": "Сначала выберите изделие из каталога."}, status=400)
        tender.cost_per_unit = tender.product.cost_price
        tender.save(update_fields=["cost_per_unit", "updated_at"])
        return Response(TenderSerializer(tender).data)

    @action(detail=True, methods=["post"])
    def make_contract(self, request, pk=None):
        """Выигранный тендер → договор (одной кнопкой, без повторного ввода)."""
        from contracts.models import Customer, Contract
        tender = self.get_object()
        if tender.status != Tender.Status.WON:
            return Response({"detail": "Договор создаётся только из выигранного тендера."}, status=400)
        if tender.contract:
            return Response({"detail": f"Договор уже создан: №{tender.contract.number}"}, status=400)

        number = request.data.get("number") or f"Т-{tender.purchase_no or tender.id}"
        if Contract.objects.filter(number=number).exists():
            return Response({"detail": f"Договор с номером {number} уже существует."}, status=400)

        customer, _ = Customer.objects.get_or_create(name=tender.customer_name.strip()[:255])
        contract = Contract.objects.create(
            number=number, customer=customer,
            title=f"{tender.item_name} — {tender.qty} шт",
            amount=tender.plan_total or tender.customer_total,
            status=Contract.Status.NEW,
            specification=tender.note,
            manager=tender.manager or request.user,
            deadline=request.data.get("deadline") or None,
        )
        tender.contract = contract
        tender.save(update_fields=["contract", "updated_at"])
        return Response(TenderSerializer(tender).data)

    @action(detail=False, methods=["get"])
    def funnel(self, request):
        """Сводка воронки: сколько лотов и денег на каждой стадии."""
        rows = {}
        for t in self.filter_queryset(self.get_queryset()):
            r = rows.setdefault(t.status, {"status": t.status,
                                           "status_display": t.get_status_display(),
                                           "count": 0, "plan_total": 0.0, "profit": 0.0})
            r["count"] += 1
            r["plan_total"] += float(t.plan_total)
            r["profit"] += float(t.profit)
        urgent = [{"id": t.id, "item": t.item_name, "customer": t.customer_name,
                   "deadline": str(t.deadline), "days_left": t.days_left}
                  for t in self.get_queryset() if t.is_urgent]
        return Response({"stages": list(rows.values()), "urgent": urgent})

    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        """Выгрузка плана закупок в формате рабочей таблицы заказчика."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "план закупок"
        ws.append(["Площадка", "Номер закупки", "Номер лота", "От какой фирмы",
                   "Организация", "Наименование товара", "Дата окончания", "Время окончания",
                   "Срок поставки", "Кол-во", "Цена заказчика", "Сумма заказчика",
                   "План цена", "Себестоимость за ед.", "Себес общий", "План сумма",
                   "Прибыль", "Маржа %", "Статус", "Решение", "Ответственный"])
        for t in self.filter_queryset(self.get_queryset()):
            ws.append([
                t.platform.name if t.platform else "", t.purchase_no, t.lot_no,
                t.own_company.name if t.own_company else "", t.customer_name, t.item_name,
                str(t.deadline or ""), str(t.deadline_time or ""), t.delivery_days,
                t.qty, float(t.price), float(t.customer_total),
                float(t.plan_price), float(t.cost_per_unit), float(t.cost_total),
                float(t.plan_total), float(t.profit), round(t.margin_percent, 1),
                t.get_status_display(), t.decision,
                t.manager.username if t.manager else "",
            ])
        buf = io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="tenders.xlsx"'
        return resp

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser])
    def import_excel(self, request):
        """Импорт плана закупок из рабочей таблицы заказчика.
        Колонки распознаются по названиям (рус.), лишние игнорируются."""
        from openpyxl import load_workbook
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "Файл не передан (поле file)."}, status=400)
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Файл пуст."}, status=400)
        header = [str(h).strip().lower() if h else "" for h in rows[0]]

        def col(*names):
            for n in names:
                for i, h in enumerate(header):
                    if n in h:
                        return i
            return None

        idx = {
            # в рабочем файле заголовок площадки бывает мусорным ("```") — берём первую колонку
            "platform": col("площадка") if col("площадка") is not None else 0,
            "purchase_no": col("номер закупки"),
            "total_sum": col("без ндс сумма"),
            "submitted": col("заявка подана"),
            "rejected": col("заявка отклонена"),
            "lot_no": col("номер лота"),
            "own": col("от какой фирмы", "с какой фирмы"),
            "customer": col("организация"),
            "item": col("наименование товара", "предмет закупки"),
            "deadline": col("дата окончание", "дата окончания"),
            "dtime": col("время окончаня", "время окончания"),
            "delivery": col("время поставки", "срок поставки"),
            "qty": col("кол-во"),
            "price": col("цена"),
            "plan_price": col("план цена"),
            "samples": col("образц"),
            "decision": col("итого решение", "решение"),
        }

        def val(row, key):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        def num(v):
            if v is None or v == "":
                return 0
            try:
                return float(str(v).replace("\xa0", "").replace(" ", "")
                             .replace("₸", "").replace(",", "."))
            except (ValueError, TypeError):
                return 0

        created, skipped, errors = 0, 0, []
        for i, row in enumerate(rows[1:], start=2):
            try:
                item = str(val(row, "item") or "").strip()
                if not item:
                    skipped += 1
                    continue
                platform = None
                pname = str(val(row, "platform") or "").strip()
                if pname and pname != "```":
                    platform, _ = Platform.objects.get_or_create(name=pname[:100])
                own = None
                oname = str(val(row, "own") or "").strip()
                if oname:
                    own, _ = OwnCompany.objects.get_or_create(name=oname[:150])

                deadline = None
                dv = val(row, "deadline")
                if isinstance(dv, datetime):
                    deadline = dv.date()
                elif isinstance(dv, str) and dv.strip():
                    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m."):
                        try:
                            d = datetime.strptime(dv.strip()[:10], fmt)
                            deadline = (d.replace(year=timezone.localdate().year).date()
                                        if fmt == "%d.%m." else d.date())
                            break
                        except ValueError:
                            continue

                # если цену не указали, но есть сумма и количество — считаем цену за единицу
                qty = int(num(val(row, "qty")))
                price = num(val(row, "price"))
                total = num(val(row, "total_sum"))
                if not price and total and qty:
                    price = round(total / qty, 2)

                # статус по отметкам в файле
                status = Tender.Status.PLANNED
                if val(row, "rejected"):
                    status = Tender.Status.REJECTED
                elif val(row, "submitted"):
                    status = Tender.Status.SUBMITTED
                if "отбой" in str(val(row, "decision") or "").lower():
                    status = Tender.Status.DECLINED

                dtime = val(row, "dtime")
                Tender.objects.create(
                    platform=platform, own_company=own,
                    purchase_no=str(val(row, "purchase_no") or "").strip()[:100],
                    lot_no=str(val(row, "lot_no") or "").strip()[:100],
                    customer_name=str(val(row, "customer") or "—").strip()[:255],
                    item_name=item[:255],
                    qty=qty, price=price, status=status,
                    plan_price=num(val(row, "plan_price")),
                    deadline=deadline,
                    deadline_time=dtime if hasattr(dtime, "hour") else None,
                    delivery_days=str(val(row, "delivery") or "").strip()[:100],
                    samples_required=str(val(row, "samples") or "").strip()[:255],
                    decision=str(val(row, "decision") or "").strip()[:255],
                )
                created += 1
            except Exception as e:  # noqa: BLE001 — сообщаем построчно, импорт не прерываем
                errors.append(f"строка {i}: {e}")
        return Response({"created": created, "skipped": skipped, "errors": errors[:20]})
